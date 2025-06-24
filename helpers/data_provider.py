# utils/data_provider.py
import json
import csv
import openpyxl  # pip install openpyxl

class DataProvider:

    @staticmethod
    def load_json(file_path):
        with open(file_path, "r", encoding="utf-8") as f:
            data = json.load(f)
        print(f"[DataProvider] Loaded JSON data from {file_path}")
        return data

    @staticmethod
    def load_csv(file_path):
        with open(file_path, newline='', encoding="utf-8") as csvfile:
            reader = csv.DictReader(csvfile)
            data = [row for row in reader]
        print(f"[DataProvider] Loaded CSV data from {file_path}")
        return data

    @staticmethod
    def load_excel(file_path, sheet_name=None):
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook[sheet_name] if sheet_name else workbook.active
        headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        data = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            data.append(dict(zip(headers, row)))
        print(f"[DataProvider] Loaded Excel data from {file_path} (Sheet: {sheet.title})")
        return data
